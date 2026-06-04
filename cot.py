import openpyxl
from openai import OpenAI

# API 配置
client = OpenAI(
    api_key='xxx',
    base_url='xxx'
)

SYSTEM_PROMPT = (
    "你是一个 SPARQL 查询生成专家。请按照以下步骤进行推理：\n"
    "1. 提取问题中的实体（例如疾病名、药物名）和关系（例如不良反应、适应症、禁忌、相互作用）。\n"
    "2. 判断该关系属于哪个 schema 类别，可选的 schema 有：\n"
    "   药物-不良反应-疾病，药物-适应症-疾病，药物-禁忌-疾病，药物-相互作用-药物\n"
    "3. 根据映射规则生成 SPARQL 查询，格式如：\n"
    "   SELECT ?x WHERE { :实体 :关系 ?x . }\n"
    "   注意：实体是从问题中提取的，关系是通过schema 类别判断得出的。\n"
    "请先输出思考过程（用自然语言），然后在最后单独一行输出 SPARQL 查询（仅输出查询语句，不要用 ``` 包裹）。"
)

def extract_sparql_from_text(text: str) -> str:
    """提取最后一行作为 SPARQL"""
    lines = text.strip().split('\n')
    while lines and not lines[-1].strip():
        lines.pop()
    if not lines:
        return ""
    return lines[-1].strip()

def generate_sparql(question: str):
    try:
        completion = client.chat.completions.create(
            model="Qwen3-8B",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": f"问题：{question}\n请按照步骤推理并生成SPARQL。"}
            ],
            temperature=0.2,
            stream=False
        )
        full_response = completion.choices[0].message.content.strip()
        # 分离思考过程和 SPARQL
        lines = full_response.split('\n')
        if len(lines) == 0:
            return "", ""
        sparql = lines[-1].strip()
        thought = '\n'.join(lines[:-1]) if len(lines) > 1 else ""
        return thought, sparql
    except Exception as e:
        return f"错误：{e}", ""

def main():
    file_path = 'cot.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    max_row = ws.max_row
    start_row = 1  # 若第一行有标题，改为2

    for row_idx in range(start_row, max_row + 1):
        cell_c = ws.cell(row=row_idx, column=2)  # C列问题
        question = cell_c.value
        if not question:
            continue

        print(f"\n===== 第 {row_idx} 行 =====")
        print(f"问题：{question}")
        thought, sparql = generate_sparql(question)
        if thought:
            print("思考过程：")
            print(thought)
        else:
            print("思考过程：无")
        print("生成的 SPARQL：")
        print(sparql)

        # 将 SPARQL 写入 F 列（第6列），E列留空或可删除原有内容
        ws.cell(row=row_idx, column=5, value=sparql)  # F列写入
        # 若希望清除E列原有内容，可取消下一行注释
        # ws.cell(row=row_idx, column=5, value=None)

        # 每行保存一次
        wb.save(file_path)

    print("\n所有问题处理完成！")

if __name__ == '__main__':
    main()