import openpyxl
from openai import OpenAI
import time

# API 配置（base_url 已修正）
client = OpenAI(
    api_key='xxx',
    base_url='xxx'
)

def ask_question(question_text):
    # 将原来的系统提示和用户问题合并到一个 user 消息中
    enhanced_prompt = (
        f"你是一名医学专业人员，请严格按照用户要求输出信息。\n\n"
        f"{question_text}\n"
        f"只给出答案，应准确全面，按1、2、3、4、5、、、等等顺序列出，"
        f"不要输出其他任何内容（如推理、解释、说明、注意事项等）。"
    )
    try:
        completion = client.chat.completions.create(
            model="Ling-2.6-1T",
            messages=[
                {'role': 'user', 'content': enhanced_prompt}
            ],
            stream=True
        )
        answer = ""
        for chunk in completion:
            if chunk.choices and len(chunk.choices) > 0:
                delta = chunk.choices[0].delta
                if delta.content:
                    answer += delta.content
        return answer.strip()
    except Exception as e:
        print(f"请求失败: {e}")
        return f"错误: {e}"

def main():
    wb = openpyxl.load_workbook('disease_questions.xlsx')
    ws = wb['mayi']
    max_row = ws.max_row
    start_row = 1   # 如果第一行是表头，请改为2

    for row_idx in range(start_row, max_row + 1):
        cell_question = ws.cell(row=row_idx, column=1)
        question = cell_question.value
        if not question:
            continue

        print(f"\n正在处理第 {row_idx} 个问题: {question}")
        answer = ask_question(question)

        ws.cell(row=row_idx, column=2, value=answer)
        print(f"答案已写入 B{row_idx}")

        wb.save('disease_questions.xlsx')
        time.sleep(1)

    print("\n所有问题处理完成！")

if __name__ == '__main__':
    main()