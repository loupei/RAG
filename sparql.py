# import openpyxl
# from SPARQLWrapper import SPARQLWrapper, JSON
#
# # GraphDB 仓库端点
# endpoint = "xxx"
# sparql = SPARQLWrapper(endpoint)
# sparql.setReturnFormat(JSON)
#
# def execute_sparql(query: str) -> str:
#     """执行 SPARQL 查询，返回结果字符串（多个值用分号分隔，单值直接返回，无结果返回“无”）"""
#     if not query or not query.strip():
#         return "无"
#     try:
#         sparql.setQuery(query)
#         results = sparql.query().convert()
#         bindings = results.get("results", {}).get("bindings", [])
#         if not bindings:
#             return "无"
#         # 提取第一个变量的所有值（通常查询只有一个变量）
#         # 获取变量名（假设第一个变量就是我们要的答案）
#         vars_list = results["head"]["vars"]
#         if not vars_list:
#             return "无"
#         var_name = vars_list[0]
#         values = [binding[var_name]["value"] for binding in bindings if var_name in binding]
#         if not values:
#             return "无"
#         return "; ".join(values)
#     except Exception as e:
#         return f"错误: {e}"
#
# def main():
#     file_path = "cot.xlsx"
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb.active
#     max_row = ws.max_row
#     start_row = 1  # 如果第一行是标题，改为2
#
#     for row_idx in range(start_row, max_row + 1):
#         cell_e = ws.cell(row=row_idx, column=5)  # E列：SPARQL查询
#         query = cell_e.value
#         if not query:
#             continue
#
#         print(f"正在执行第 {row_idx} 行的查询...")
#         result = execute_sparql(str(query))
#
#         # 写入 F 列
#         ws.cell(row=row_idx, column=6, value=result)
#         print(f"结果已写入 F{row_idx}")
#
#         # 每行保存一次，防止数据丢失
#         wb.save(file_path)
#
#     print("所有查询执行完成！")
#
# if __name__ == "__main__":
#     main()


import openpyxl
from openai import OpenAI

# API 配置
client = OpenAI(
    api_key='xxx',
    base_url='xxx'
)

def generate_sparql(question: str, hint: str) -> str:
    """调用大模型，根据问题和提示生成 SPARQL 查询，返回查询字符串。"""
    system_msg = {
        'role': 'system',
        'content': (
            "你是一名 SPARQL 查询生成专家。请根据用户提供的图谱模式（Schema，药物-不良反应-疾病，药物-适应症-疾病，药物-禁忌-疾病，药物-相互作用-药物。）和问题，"
            "给出一个查询示例，PREFIX : <http://rag.org/> SELECT ?x WHERE {:氟尿嘧啶 :不良反应 ?x .}"
            "只输出 SPARQL 查询语句，不要输出任何解释、注释或额外文字，按照Schema给出查询，查询中的关系写为中文。"
        )
    }
    user_msg = {
        'role': 'user',
        'content': f"问题：{question}\n\n提示：{hint}\n\n请生成 SPARQL 查询。"
    }
    try:
        completion = client.chat.completions.create(
            model="Qwen3-8B",
            messages=[system_msg, user_msg],
            stream=False,
            temperature=0.1  # 低温度，提高确定性和格式规范性
        )
        return completion.choices[0].message.content.strip()
    except Exception as e:
        print(f"API 请求失败：{e}")
        return f"错误：{e}"

def main():
    file_path = 'cot.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # 假设只有一个工作表，也可指定 wb['Sheet1']
    max_row = ws.max_row
    start_row = 1  # 如果第一行是标题，从第2行开始

    for row_idx in range(start_row, max_row + 1):
        cell_q = ws.cell(row=row_idx, column=2)  # C列问题
        cell_h = ws.cell(row=row_idx, column=3)  # D列提示
        question = cell_q.value
        hint = cell_h.value if cell_h.value else ''

        if not question:
            continue

        print(f"正在处理第 {row_idx} 行，问题：{question[:50]}...")
        sparql = generate_sparql(question, hint)

        # 将生成的 SPARQL 写入 E 列
        ws.cell(row=row_idx, column=4, value=sparql)  # E列
        print(f"SPARQL 已写入 E{row_idx}")

        # 每处理一行保存一次，防止中断丢失数据
        wb.save(file_path)

    print("\n所有问题处理完成！")

if __name__ == '__main__':
    main()