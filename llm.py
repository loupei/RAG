import openpyxl
from openai import OpenAI
import time

# API 配置（使用你提供的参数）
client = OpenAI(
    api_key='xxx',
    base_url='xxx'
)

def ask_question(question_text):
    """
    调用大模型（流式），返回完整答案字符串。
    强制要求模型只输出药物引起的不良反应，避免多余解释。
    """
    # 在原始问题后追加强约束
    enhanced_prompt = f"{question_text}\n只给出答案，应准确全面，在全面的基础上归纳其影响的是什么系统、有什么表现，按123点列出，不要输出其他任何内容（如推理、解释、说明、注意事项等）。"
    try:
        completion = client.chat.completions.create(
            model="Qwen3-8B",
            messages=[
                {'role': 'system', 'content': '你是一名医学专业人员，请严格按照用户要求输出信息。'},
                {'role': 'user', 'content': enhanced_prompt}
            ],
            stream=True
        )
        answer = ""
        for chunk in completion:
            if chunk.choices and len(chunk.choices) > 0:
                delta = chunk.choices[0].delta
                if delta.content:
                    answer += delta.content
        return answer.strip()
    except Exception as e:
        print(f"请求失败: {e}")
        return f"错误: {e}"

def main():
    # 打开问题文件
    wb = openpyxl.load_workbook('drug_questions.xlsx')
    ws = wb.active  # 假设问题在第一个工作表

    # 确定最大行数（A列有内容）
    max_row = ws.max_row
    start_row = 1  # 如果第一行没有表头，从第1行开始；若有表头可改为2

    for row_idx in range(start_row, max_row + 1):
        cell_question = ws.cell(row=row_idx, column=1)
        question = cell_question.value
        if not question:
            continue

        print(f"\n正在处理第 {row_idx} 个问题: {question}")
        answer = ask_question(question)

        # 写入 B 列
        ws.cell(row=row_idx, column=2, value=answer)
        print(f"答案已写入 B{row_idx}")

        # 立即保存，防止程序中断丢失数据
        wb.save('drug_questions.xlsx')
        # 适当延迟，避免请求过快（根据 API 限制调整）
        time.sleep(1)

    print("\n所有问题处理完成！")

if __name__ == '__main__':
    main()