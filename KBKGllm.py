import openpyxl
from openai import OpenAI
import time

# API 配置（使用你提供的参数）
client = OpenAI(
    api_key='xxx',
    base_url='xxx'
)

def ask_with_prompt(question_text, prompt_text):
    """
    根据问题和提示，调用大模型（流式），返回完整答案。
    提示作为 system 指令，强制模型严格遵守提示回复。
    """
    if prompt_text and prompt_text.strip():
        system_msg = {
            'role': 'system',
            'content': f"你是一名医学专业人员。请严格遵循以下提示来回答问题，除非提示为空否则也不要使用自己的知识，只使用提示中的信息回答，不要输出任何多余内容（如解释、推理、注意事项等），对提示进行总结，使用非结构化自然语言回答，做好归纳总结。\n提示：{prompt_text}"
        }
    else:
        system_msg = {'role': 'system', 'content': '你是一名医学专业人员，请严格按照用户要求输出信息。'}

    try:
        completion = client.chat.completions.create(
            model="Qwen3-8B",
            messages=[
                system_msg,
                {'role': 'user', 'content': question_text}
            ],
            stream=True
        )
        answer = ""
        for chunk in completion:
            if chunk.choices and len(chunk.choices) > 0:
                delta = chunk.choices[0].delta
                if delta.content:
                    answer += delta.content
        return answer.strip()
    except Exception as e:
        print(f"请求失败: {e}")
        return f"错误: {e}"

def main():
    # 打开问题文件，并指定工作表 'kbkgllm'
    wb = openpyxl.load_workbook('disease_questions.xlsx')
    ws = wb['kbkgllm']   # 修改点1：使用指定的工作表，而非 active

    # 确定最大行数（A列有内容）
    max_row = ws.max_row
    start_row = 1  # 如果第一行没有表头，从第1行开始；若有表头可改为2

    for row_idx in range(start_row, max_row + 1):
        cell_question = ws.cell(row=row_idx, column=1)   # A列：问题
        cell_prompt = ws.cell(row=row_idx, column=2)     # B列：提示
        question = cell_question.value
        prompt = cell_prompt.value if cell_prompt.value else ""

        if not question:
            continue

        print(f"\n正在处理第 {row_idx} 个问题: {question}")
        print(f"提示: {prompt}")
        answer = ask_with_prompt(question, prompt)

        # 答案写入 C 列
        ws.cell(row=row_idx, column=3, value=answer)
        print(f"答案已写入 C{row_idx}")

        # 立即保存
        wb.save('disease_questions.xlsx')
        time.sleep(1)

    print("\n所有问题处理完成！")

if __name__ == '__main__':
    main()